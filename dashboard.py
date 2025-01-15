import customtkinter as ctk
from tkinter import messagebox, ttk  # Importando ttk para criar a tabela
from tkcalendar import DateEntry
from database import Database
from excel_handler import ExcelHandler
import os
import pandas as pd  # Importando pandas para manipular dados do Excel
from openpyxl import load_workbook
import sqlite3

class DashboardEmpresarial:
    def __init__(self, root):
        self.root = root
        self.root.title("ERP Empresarial")
        self.root.geometry("1200x700")
        ctk.set_appearance_mode("dark")
        ctk.set_default_color_theme("blue")

        self.database = Database("clientes.db")
        self.excel_handler = ExcelHandler("clientes.xlsx")

        self.create_widgets()

    def create_widgets(self):
        self.create_button_frame()
        self.frames = {}
        self.main_frame = ctk.CTkFrame(self.root)
        self.main_frame.pack(fill=ctk.BOTH, expand=True)

        self.init_home_tab()
        self.init_clientes_tab()
        self.init_produtos_tab()  # Adiciona a aba Produtos
        self.init_vendas_tab()     # Adiciona a aba Vendas
        self.init_relatorios_tab()  # Adiciona a aba Relatórios

        self.switch_tab(1)  # Abre diretamente na aba "Clientes"

    def create_button_frame(self):
        button_frame = ctk.CTkFrame(self.root)
        button_frame.pack(fill=ctk.X, pady=10)

        buttons = ["HOME", "Clientes", "Produtos", "Vendas", "Relatórios"]
        for i, name in enumerate(buttons):
            button = ctk.CTkButton(
                button_frame,
                text=name,
                command=lambda i=i: self.switch_tab(i),
                corner_radius=8,
                fg_color="#1f6aa5",
                hover_color="#144870"
            )
            button.pack(side=ctk.LEFT, padx=5)

    def switch_tab(self, index):
        for frame in self.frames.values():
            frame.pack_forget()
        self.frames[index].pack(fill=ctk.BOTH, expand=True)

    def init_home_tab(self):
        frame = ctk.CTkFrame(self.main_frame)
        self.frames[0] = frame

        ctk.CTkLabel(frame, text="Pesquisar Cliente", font=("Arial", 24, "bold")).pack(pady=10)
        self.pesquisaInput = ctk.CTkEntry(frame, width=400, font=("Arial", 16))
        self.pesquisaInput.pack(pady=5)

        ctk.CTkButton(frame, text="Pesquisar", command=self.pesquisar_cliente, font=("Arial", 16)).pack(pady=5)

        self.clienteInfo = ctk.CTkTextbox(frame, width=800, height=500, state=ctk.DISABLED, font=("Arial", 18))
        self.clienteInfo.pack(pady=10)

        ctk.CTkButton(frame, text="Copiar Informações", command=self.copiar_informacoes, font=("Arial", 16)).pack(pady=5)

    def init_clientes_tab(self):
        frame = ctk.CTkFrame(self.main_frame)
        self.frames[1] = frame

        ctk.CTkLabel(frame, text="Cadastro de Clientes", font=("Arial", 20, "bold")).pack(pady=10)

        form_frame = ctk.CTkFrame(frame)
        form_frame.pack(pady=20, padx=20, fill=ctk.BOTH, expand=True)

        labels = ["Nome", "CPF", "RG", "Sexo", "Telefone", "Endereço", "PIS/NIS", "NIP", "CEI", "RGP", "Email", "Titulo de Eleitor","Data de Nascimento","Data Inicio Atividade"]
        self.inputs = {}

        for i, label in enumerate(labels):
            self.create_input_field(form_frame, label, i)

        add_button = ctk.CTkButton(frame, text="Adicionar Cliente", command=self.add_cliente, width=250, height=40)
        add_button.pack(pady=20)

    def init_produtos_tab(self):
        frame = ctk.CTkFrame(self.main_frame)
        self.frames[2] = frame

        ctk.CTkLabel(frame, text="Produtos", font=("Arial", 20, "bold")).pack(pady=10)
        ctk.CTkLabel(frame, text="Esta funcionalidade está em criação.", font=("Arial", 16), text_color="red").pack(pady=20)

    def init_vendas_tab(self):
        frame = ctk.CTkFrame(self.main_frame)
        self.frames[3] = frame

        ctk.CTkLabel(frame, text="Vendas", font=("Arial", 20, "bold")).pack(pady=10)
        ctk.CTkLabel(frame, text="Esta funcionalidade está em criação.", font=("Arial", 16), text_color="red").pack(pady=20)

    def init_relatorios_tab(self):
        frame = ctk.CTkFrame(self.main_frame)
        self.frames[4] = frame

        ctk.CTkLabel(frame, text="Relatórios", font=("Arial", 20, "bold")).pack(pady=10)
        ctk.CTkLabel(frame, text="Clique no botão abaixo para carregar os dados do Excel.", font=("Arial", 16)).pack(
            pady=20)

        # Frame para tabela e scrollbar
        table_frame = ctk.CTkFrame(frame)
        table_frame.pack(pady=10, fill=ctk.BOTH, expand=True)

        # Criação da tabela
        self.tree = ttk.Treeview(table_frame, columns=(
            "Nome", "CPF", "RG", "Sexo", "Telefone", "Endereço", "PIS/NIS", "NIP", "CEI", "RGP", "Email", "Titulo de Eleitor","Data de Nascimento","Data Inicio Atividade"), show="headings")

        # Configuração das colunas
        for col in self.tree["columns"]:
            self.tree.heading(col, text=col)
            self.tree.column(col, anchor="center", width=150)  # Largura padrão para as colunas

        # Adicionar barras de rolagem
        scroll_x = ttk.Scrollbar(table_frame, orient="horizontal", command=self.tree.xview)
        scroll_y = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(xscrollcommand=scroll_x.set, yscrollcommand=scroll_y.set)

        # Posicionamento da tabela e barras de rolagem
        self.tree.grid(row=0, column=0, sticky="nsew")
        scroll_x.grid(row=1, column=0, sticky="ew")
        scroll_y.grid(row=0, column=1, sticky="ns")

        # Configuração de layout responsivo
        table_frame.rowconfigure(0, weight=1)
        table_frame.columnconfigure(0, weight=1)

        # Botão para carregar os dados
        load_button = ctk.CTkButton(frame, text="Carregar Dados do Excel", command=self.carregar_dados_excel,
                                    font=("Arial", 16))
        load_button.pack(pady=5)

    def carregar_dados_excel(self):
        file_path = "clientes.xlsx"
        if os.path.exists(file_path):
            # Lê os dados do Excel usando pandas
            try:
                df = pd.read_excel(file_path)
                for i in self.tree.get_children():  # Limpa a tabela antes de carregar novos dados
                    self.tree.delete(i)

                for _, row in df.iterrows():
                    self.tree.insert("", "end", values=row.tolist())  # Insere cada linha na tabela
            except Exception as e:
                messagebox.showerror("Erro", f"Erro ao ler o arquivo Excel: {e}")
        else:
            messagebox.showwarning("Erro", "O arquivo 'clientes.xlsx' não foi encontrado.")

    def create_input_field(self, form_frame, label, index):
        row = index // 2
        column = index % 2

        ctk.CTkLabel(form_frame, text=label, font=("Arial", 14)).grid(row=row, column=column * 2, sticky=ctk.W, pady=10, padx=10)

        if label == "Data de Nascimento" or label == "Data Inicio Atividade":
            input_field = DateEntry(form_frame, width=20, date_pattern="yyyy-mm-dd", borderwidth=2)
        elif label == "Sexo":
            input_field = ctk.CTkComboBox(form_frame, values=["Masculino", "Feminino"], width=350, height=35)
            input_field.grid(row=row, column=column * 2 + 1, pady=15, padx=10, sticky=ctk.E)  # Alinhado à direita
        else:
            input_field = ctk.CTkEntry(form_frame, width=400, height=35)

        input_field.grid(row=row, column=column * 2 + 1, pady=15, padx=10, sticky=ctk.W)
        self.inputs[label] = input_field

    def pesquisar_cliente(self):
        pesquisa = self.pesquisaInput.get()
        if pesquisa:
            try:
                if os.path.exists("clientes.xlsx"):
                    self._search_cliente_in_excel(pesquisa)
                else:
                    messagebox.showwarning("Erro", "O arquivo 'clientes.xlsx' não foi encontrado.")
            except Exception as e:
                messagebox.showerror("Erro", f"Ocorreu um erro ao buscar os dados: {e}")
        else:
            messagebox.showwarning("Erro", "Digite um nome para pesquisar!")

    def _search_cliente_in_excel(self, pesquisa):
        workbook = load_workbook("clientes.xlsx")
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            nome = row[0]
            if pesquisa.lower() in nome.lower():
                self._display_cliente_info(row)
                return
        messagebox.showwarning("Resultado da Pesquisa", "Nenhum cliente encontrado!")

    def _display_cliente_info(self, cliente):
        if len(cliente) == 14:
            self.clienteInfo.configure(state=ctk.NORMAL)
            self.clienteInfo.delete("1.0", ctk.END)
            self.clienteInfo.insert(ctk.END, self.format_cliente_info(cliente))
            self.clienteInfo.configure(state=ctk.DISABLED)
        else:
            messagebox.showwarning("Erro", f"Dados incompletos para o cliente {cliente[0]}. Verifique o arquivo.")

    def format_cliente_info(self, cliente):
        return f"""
        Nome: {cliente[0]}, 

        CPF: {cliente[1]}

        RG: {cliente[2]}

        Sexo: {cliente[3]}

        Telefone: {cliente[4]}

        Endereço: {cliente[5]}

        PIS/NIS: {cliente[6]}

        NIP: {cliente[7]}

        CEI: {cliente[8]}

        RGP: {cliente[9]}

        Email: {cliente[10]}

        Título de Eleitor: {cliente[11]}

        Data de Nascimento: {cliente[12]}

        Data Inicio Atividade: {cliente[13]}

        """

    def copiar_informacoes(self):
        self.root.clipboard_clear()
        self.root.clipboard_append(self.clienteInfo.get("1.0", ctk.END).strip())
        self.root.update()
        messagebox.showinfo("Sucesso", "Informações copiadas para a área de transferência!")

    def add_cliente(self):
        try:
            values = self._get_cliente_data()
            if all(values.values()):
                self.database.insert_cliente(tuple(values.values()))
                self.excel_handler.exportar_cliente(tuple(values.values()))
                messagebox.showinfo("Sucesso", "Cliente adicionado com sucesso!")
                self._clear_input_fields()
            else:
                messagebox.showwarning("Erro", "Preencha todos os campos!")
        except sqlite3.Error as e:
            messagebox.showerror("Erro no Banco de Dados", f"Ocorreu um erro ao salvar no banco de dados:\n{e}")
        except Exception as e:
            messagebox.showerror("Erro", f"Ocorreu um erro inesperado:\n{e}")

    def _get_cliente_data(self):
        return {label: input_field.get() for label, input_field in self.inputs.items()}

    def _clear_input_fields(self):
        for label, input_field in self.inputs.items():
            if isinstance(input_field, ctk.CTkComboBox):  # Verifica se é um CTkComboBox
                input_field.set("")  # Redefine o valor do ComboBox para vazio
            else:
                input_field.delete(0, ctk.END)  # Limpa campos de entrada (Entry ou CTkEntry)


if __name__ == "__main__":
    root = ctk.CTk()
    app = DashboardEmpresarial(root)
    root.mainloop()
