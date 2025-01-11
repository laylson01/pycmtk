import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from tkinter import messagebox
import sqlite3

class ExcelHandler:
    def __init__(self, filename):
        self.filename = filename

    def exportar_cliente(self, cliente_data):
        workbook = self._load_workbook()
        sheet = workbook.active
        sheet.append(cliente_data)
        self._format_sheet(sheet)
        workbook.save(self.filename)
        messagebox.showinfo("Sucesso", "Dados exportados para Excel com sucesso.")

    def _load_workbook(self):
        if os.path.exists(self.filename):
            return load_workbook(self.filename)
        else:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Clientes"
            headers = ["Nome", "CPF", "RG", "Data de Nascimento", "Sexo", "Telefone", "Endereço", "PIS/NIS", "NIP", "CEI", "RGP", "Email", "Data Inicio Atividade", "Titulo de Eleitor"]
            sheet.append(headers)
            self._format_sheet(sheet)
            return workbook

    def _format_sheet(self, sheet):
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
